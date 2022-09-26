#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Wed Sep 21 15:43:58 2022

@author: carlosswanton
"""

# Import libraries: 

import pandas as pd 
import numpy as np 
import os 
import openpyxl
from openpyxl import Workbook, load_workbook 

# Get username
tm = input("Cual es el nuevo periodo de la bateria? (Ej: 20220930: ")

user = os.getlogin()
if user == "root": 
    indicadores_sociales_dir = ("/Users/carlosswanton/Desktop/tabulados_enemdu/")
    sector_petrolero_dir = ("/Users/carlosswanton/Desktop/petroleo/")
    presupuesto_dir = ("/Users/carlosswanton/Desktop/")
    balanza_comercial_dir = ("/Users/carlosswanton/Desktop/balanza/")
    bateria_dir = ("/Volumes/GoogleDrive/.shortcut-targets-by-id/1crJlLLDGbhJqLkDtjyKBhWLOCUY5y9Tv/01 IE/06 Bateria estadística/")
    bateria_nueva = bateria_dir + tm + "/" + "Batería estadística IE " + tm[0:6] + ".xlsx"
    
"Sección Indicadores Sociales"

#Importar Excel del tabulado mensual 
# El periodo de enemdu es un mes antes de la bateria. 
enemdu_periodo = input("Especificar el período de indicadores laborales (Ej: 202208): ")

enemdu = pd.read_excel(indicadores_sociales_dir + enemdu_periodo + ".xlsx", sheet_name = "2. Tasas", 
                       header = None, names = ["date", "indicadores", "nacional", "urbano"],
                       usecols = "B:E") 

enemdu1 = enemdu.tail(14) #It should always be the last 14. Estructura del boletín.
#You can also keep duplicates last occ.
# Fix y calcular.

enemdu1["indicadores"].replace({"Empleo Bruto (%)" : "empleo_bruto", "Empleo Global (%)" : "empleo_global",
                               "Subempleo (%)" : "subempleo",
                               "Empleo Adecuado/Pleno (%)" :"empleo_adecuado", "Subempleo por  insuficiencia de ingresos (%)":
                              "subempleo_por_ingresos", "Empleo no Remunerado (%)" : "empleo_no_remunerado", 
                              "Otro Empleo no pleno(%)":"otro_empleo_no_pleno", "Empleo no Clasificado (%)": "empleo_no_clasi",
                              "Desempleo (%)" : "desempleo", "Desempleo Abierto (%)": "desempleo_abierto", 
                              "Desempleo Oculto (%)": "desempleo_oculto","Participación Global (%)": "participación_global",
                              "Participación Bruta (%)" : "participación_bruta",
                              "Subempleo por insuficiencia de tiempo de trabajo (%)":"subempleo_tiempo"}, inplace = True)
# Convertir a porcentajes:
enemdu1["nacional"] = enemdu1["nacional"] / 100 
enemdu1["urbano"] = enemdu1["urbano"] / 100 

# Set index:
enemdu1.set_index(enemdu1["indicadores"], drop = True, inplace = True)
enemdu1.drop(columns = ["indicadores"], inplace = True)

# Variables Empleo: 

empleo_adecuado_nacional = enemdu1.loc["empleo_adecuado", "nacional"]
empleo_adecuado_urbano = enemdu1.loc["empleo_adecuado", "urbano"]
desempleo_nacional = enemdu1.loc["desempleo", "nacional"]
desempleo_urbano = enemdu1.loc["desempleo", "urbano"]

a = enemdu1.loc["subempleo", "nacional"] 
b = enemdu1.loc["empleo_no_remunerado", "nacional"] 
c = enemdu1.loc["otro_empleo_no_pleno", "nacional"]
empleo_inadecuado_nacional = a+b+c

a = enemdu1.loc["subempleo", "urbano"] 
b = enemdu1.loc["empleo_no_remunerado", "urbano"] 
c = enemdu1.loc["otro_empleo_no_pleno", "urbano"]
empleo_inadecuado_urbano = a+b+c


"Obtener Poblacion Economicamente Activa"

pea = pd.read_excel(indicadores_sociales_dir + enemdu_periodo + ".xlsx", sheet_name = "1. Poblaciones", 
                       header = None, names = ["date", "indicadores", "nacional"],
                       usecols = "B:D") 

pea.drop_duplicates(subset = "indicadores", keep = "last", inplace = True)

pea["indicadores"].replace({"Población Económicamente Activa" : "PEA"}, inplace = True)
filt = (pea["indicadores"] == "PEA")
x = pea.loc[filt]

x.reset_index(drop = True, inplace = True)
poblacion = x.loc[0, "nacional"]

poblacion = poblacion / 1000000

"Costo de Canasta Basica e Ingreso Familiar"

canasta = load_workbook(indicadores_sociales_dir + enemdu_periodo + "_CB" + ".xlsx")
costos = canasta["1. NACIONAL"]
# Variables Canasta Familiar:
    
costo = costos["D16"].value
ingreso_familiar = costos["E16"].value

if costo < ingreso_familiar:
    restriccion_en_consumo = 0 
else: 
    restriccion_en_consumo = (ingreso_familiar - costo) / costo

"Seccion Sector Petrolero"

# Importar petróleo consolidado para obtener producción petrolera:
p411 = pd.read_excel(sector_petrolero_dir + "Petroleo-conso.xlsx", 
                     sheet_name = "p411", header = None, names = ["date", "total", "promedio_diario","público", "privado"],
                     usecols = "C:G", skiprows = 279) #cambiar una vez al ano.

# Crear número de días:
p411["days"] = p411["date"].dt.days_in_month
# Variables de produccion petrolera: 

producción_nacional_anual = (p411["total"].sum() / 1000)
producción_nacional_diaria = (p411["total"].sum() / p411["days"].sum())
producción_pública_diaria = (p411["público"].sum() /  p411["days"].sum())
producción_privada_diaria = (p411["privado"].sum() /  p411["days"].sum())

#Obtener el precio promedio del crudo de Oriente:

precio_petroleo = pd.read_excel(sector_petrolero_dir + "Petroleo-conso.xlsx", 
                     sheet_name = "p412b", header = None, names = ["date", "Oriente"],
                     usecols = "C:D", skipfooter = 8) # que se mantenga los 8.

precio_petroleo.dropna(inplace=True)
precio_petroleo = precio_petroleo.tail(1)
precio_petroleo.reset_index(drop = True, inplace = True)

precio_promedio_crudo_oriente = precio_petroleo.loc[0,"Oriente"]

"Presupuesto General del Estado"
"VERIFICAR QUE ESTEN DOS ANOS EN COMPARACION"

presupuesto = load_workbook(presupuesto_dir + "Presupuesto Mensual.xlsx")
td = presupuesto["TD"]

# Variables presupuesto 

ingresos_totales = td["D8"].value
ingresos_totales = ingresos_totales / 1000000

ingresos_tributarios = td["D9"].value
ingresos_tributarios = ingresos_tributarios / 1000000

ingresos_petroleros = td["D12"].value
ingresos_petroleros = ingresos_petroleros / 1000000

gastos_totales = td["D17"].value
gastos_totales = gastos_totales / 1000000

gastos_corrientes = td["D18"].value
gastos_corrientes = gastos_corrientes / 1000000

sueldos_y_salarios = td["D20"].value
sueldos_y_salarios = sueldos_y_salarios / 1000000

gastos_de_inversión = td["D27"].value
gastos_de_inversión = gastos_de_inversión / 1000000

resultado_ingresos_gastos = ingresos_totales - gastos_totales

"Balanza Comercial (Sector Externo)"
# el periodo de la balanza comercial son dos meses antes de la bateria.
balanza_periodo = input("Especificar el período - Balanza (Ej: 202207): ")

comex = load_workbook(balanza_comercial_dir + balanza_periodo + ".xlsx")
balanza = comex.active
print(balanza)

# Los valores de la balanza tienen que estar en la misma columna G. Verificar.

balanza_comercial_total = balanza["G21"].value

balanza_comercial_petrolera = balanza["G23"].value

exportaciones_petroleras = balanza["G24"].value

importaciones_petroleras = balanza["G25"].value

balanza_comercial_no_petrolera = balanza["G26"].value

exportaciones_no_petroleras = balanza["G27"].value

importaciones_no_petroleras = balanza["G28"].value


"Elaboracion de la bateria"

# Importar la ultima bateria y actualizarla con la version actual.

excel = load_workbook("/Users/carlosswanton/Desktop/Batería/202209/Batería estadística IE 202209.xlsx")
bateria = excel.active


bateria["Q16"] = producción_nacional_anual
bateria["Q17"] = producción_nacional_diaria
bateria["Q18"] = producción_pública_diaria
bateria["Q19"] = producción_privada_diaria
bateria["Q20"] = precio_promedio_crudo_oriente

bateria["Q22"] = ingresos_totales
bateria["Q23"] = ingresos_tributarios
bateria["Q24"] = ingresos_petroleros
bateria["Q25"] = gastos_totales
bateria["Q26"] = gastos_corrientes
bateria["Q27"] = sueldos_y_salarios
bateria["Q28"] = gastos_de_inversión
bateria["Q29"] = resultado_ingresos_gastos


bateria["Q42"] = balanza_comercial_total
bateria["Q43"] = balanza_comercial_petrolera
bateria["Q44"] = exportaciones_petroleras
bateria["Q45"] = importaciones_petroleras
bateria["Q46"] = balanza_comercial_no_petrolera
bateria["Q47"] = exportaciones_no_petroleras
bateria["Q48"] = importaciones_no_petroleras


bateria["Q53"] = costo
bateria["Q54"] = ingreso_familiar
bateria["Q55"] = restriccion_en_consumo


bateria["Q57"] = poblacion
bateria["Q58"] = empleo_adecuado_nacional
bateria["Q59"] = empleo_inadecuado_nacional
bateria["Q60"] = desempleo_nacional
bateria["Q61"] = empleo_adecuado_urbano
bateria["Q62"] = empleo_inadecuado_urbano
bateria["Q63"] = desempleo_urbano

excel.save("/Users/carlosswanton/Desktop/Batería/202209/Batería estadística IE 202209.xlsx")




    
    